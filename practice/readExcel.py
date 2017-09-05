from openpyxl import Workbook
from openpyxl.compat import range
#不确定 openpyxl的版本使用这段代码，如果openpyxl.cell出现引入错误，则使用openpyxl.utils
try:
	from openpyxl.cell import get_column_letter
except ImportError:
	from openpyxl.utils import get_column_letter

wb = Workbook()
dest_filename = 'empty_book2.xlsx'
ws1 = wb.active    #第一个表
ws1.title = "range names" #第一个表名

#遍历第一个表的1到40行，赋值一个60内的随机数
for row in range(1,40):
	ws1.append(range(60))
ws2 = wb.create_sheet(title="Pi") #第二个标签
ws2['F5'] = 3.14  #在F5的位置赋值3.14
ws3 = wb.create_sheet(title="Data") #第三个标签
for row in range(10,20):  #在第三个标签的第10行27列开始复制英文字母
	for col in range(27,54):
		_ = ws3.cell(column=col,row=row,value="%s" % get_column_letter(col))
wb.save(filename=dest_filename)